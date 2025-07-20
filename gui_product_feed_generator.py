import os
import csv
import zipfile
import logging
import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
import paramiko
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

# --- GUI Variables ---
PRODUCTS_DIR = ""
OUTPUT_XML = 'google_product_feed.xml'
ZIP_NAME = 'product_feed.zip'
LOG_FILE = 'log.txt'
IMAGE_URL_PREFIX = 'https://yourdomain.com/images/'

REQUIRED_FIELDS = {'sku', 'name', 'price', 'quantity', 'shipping'}

# --- Logging Setup ---
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# --- Feed Generation Functions ---
def prettify(elem):
    rough_string = ET.tostring(elem, 'utf-8')
    return minidom.parseString(rough_string).toprettyxml(indent="  ")

def read_product_file(filepath):
    ext = filepath.lower().split('.')[-1]
    data = []
    if ext in ['csv', 'txt']:
        with open(filepath, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            data.extend(reader)
    elif ext == 'xlsx':
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: str(row[i]) if row[i] is not None else '' for i in range(len(headers))}
            data.append(item)
    return data

def create_product_item(data, image_urls):
    item = ET.Element('item')
    def add(tag, value):
        if value:
            ET.SubElement(item, tag).text = value
    def g(tag): return f'g:{tag}'

    add(g('id'), data.get('sku'))
    add(g('title'), data.get('name'))
    add(g('description'), data.get('name'))
    add(g('link'), f"https://yourdomain.com/product/{data.get('sku')}")
    add(g('price'), f"{data.get('price')} USD")
    add(g('availability'), "in stock" if int(data.get('quantity', 0)) > 0 else "out of stock")
    add(g('condition'), data.get('condition', 'new'))
    add(g('brand'), data.get('brand'))
    add(g('gtin'), data.get('gtin'))
    add(g('mpn'), data.get('mpn'))

    for i, url in enumerate(image_urls):
        if i == 0:
            add(g('image_link'), url)
        else:
            ET.SubElement(item, g('additional_image_link')).text = url

    shipping = ET.SubElement(item, g('shipping'))
    ET.SubElement(shipping, g('price')).text = f"{data.get('shipping')} USD"
    return item

def generate_xml_feed(products_dir, log_box):
    rss = ET.Element('rss', attrib={
        'version': '2.0',
        'xmlns:g': 'http://base.google.com/ns/1.0'
    })
    channel = ET.SubElement(rss, 'channel')
    ET.SubElement(channel, 'title').text = 'My Product Feed'
    ET.SubElement(channel, 'link').text = 'https://yourdomain.com'
    ET.SubElement(channel, 'description').text = 'Product feed for Google Merchant Center'

    for folder in os.listdir(products_dir):
        product_path = os.path.join(products_dir, folder)
        if not os.path.isdir(product_path):
            continue

        data_file = None
        image_files = []

        for f in os.listdir(product_path):
            if f.lower().endswith(('.csv', '.txt', '.xlsx')):
                data_file = os.path.join(product_path, f)
            elif f.lower().endswith('.jpg'):
                image_files.append(f)

        if not data_file or not image_files:
            msg = f"Skipping '{folder}': Missing data or images."
            logging.warning(msg)
            log_box.insert(tk.END, msg + '\n')
            continue

        try:
            product_data_list = read_product_file(data_file)
            for product_data in product_data_list:
                if not REQUIRED_FIELDS.issubset(product_data.keys()):
                    msg = f"Skipping '{folder}': Missing required fields."
                    logging.warning(msg)
                    log_box.insert(tk.END, msg + '\n')
                    continue

                image_urls = [f"{IMAGE_URL_PREFIX}{folder}/{img}" for img in sorted(image_files)]
                item = create_product_item(product_data, image_urls)
                channel.append(item)
        except Exception as e:
            msg = f"Error processing '{folder}': {str(e)}"
            logging.error(msg)
            log_box.insert(tk.END, msg + '\n')

    xml_str = prettify(rss)
    with open(OUTPUT_XML, 'w', encoding='utf-8') as f:
        f.write(xml_str)

    log_box.insert(tk.END, f"✅ Feed created: {OUTPUT_XML}\n")

def zip_feed_file(log_box):
    with zipfile.ZipFile(ZIP_NAME, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(OUTPUT_XML)
    log_box.insert(tk.END, f"✅ Feed zipped: {ZIP_NAME}\n")

def upload_to_sftp(host, port, user, password, remote_path, log_box):
    try:
        transport = paramiko.Transport((host, int(port)))
        transport.connect(username=user, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        remote_file = os.path.join(remote_path, ZIP_NAME)
        sftp.put(ZIP_NAME, remote_file)
        sftp.close()
        transport.close()
        log_box.insert(tk.END, f"✅ Uploaded to SFTP: {remote_file}\n")
    except Exception as e:
        msg = f"SFTP upload failed: {e}"
        logging.error(msg)
        log_box.insert(tk.END, f"❌ {msg}\n")

# --- GUI Setup ---
def launch_gui():
    def choose_folder():
        folder = filedialog.askdirectory()
        if folder:
            folder_entry.delete(0, tk.END)
            folder_entry.insert(0, folder)

    def run_pipeline(zip_it=False, sftp_it=False):
        log_box.delete('1.0', tk.END)
        folder = folder_entry.get()
        if not folder:
            messagebox.showerror("Error", "Select a products folder.")
            return
        generate_xml_feed(folder, log_box)
        if zip_it:
            zip_feed_file(log_box)
        if sftp_it:
            upload_to_sftp(
                host_entry.get(),
                port_entry.get(),
                user_entry.get(),
                pass_entry.get(),
                path_entry.get(),
                log_box
            )

    root = tk.Tk()
    root.title("Google Product Feed Generator")

    # Layout
    tk.Label(root, text="Products Folder:").grid(row=0, column=0, sticky='w')
    folder_entry = tk.Entry(root, width=50)
    folder_entry.grid(row=0, column=1)
    tk.Button(root, text="Browse", command=choose_folder).grid(row=0, column=2)

    # SFTP Config
    tk.Label(root, text="SFTP Host:").grid(row=1, column=0, sticky='w')
    host_entry = tk.Entry(root, width=30)
    host_entry.insert(0, "sftp.example.com")
    host_entry.grid(row=1, column=1)

    tk.Label(root, text="Port:").grid(row=1, column=2, sticky='e')
    port_entry = tk.Entry(root, width=5)
    port_entry.insert(0, "22")
    port_entry.grid(row=1, column=3)

    tk.Label(root, text="Username:").grid(row=2, column=0, sticky='w')
    user_entry = tk.Entry(root, width=30)
    user_entry.insert(0, "your_username")
    user_entry.grid(row=2, column=1)

    tk.Label(root, text="Password:").grid(row=2, column=2, sticky='e')
    pass_entry = tk.Entry(root, width=20, show="*")
    pass_entry.insert(0, "your_password")
    pass_entry.grid(row=2, column=3)

    tk.Label(root, text="Remote Path:").grid(row=3, column=0, sticky='w')
    path_entry = tk.Entry(root, width=50)
    path_entry.insert(0, "/upload/path/")
    path_entry.grid(row=3, column=1, columnspan=2)

    # Buttons
    tk.Button(root, text="Run Feed Only", command=lambda: run_pipeline()).grid(row=4, column=0, pady=10)
    tk.Button(root, text="Run Feed + ZIP", command=lambda: run_pipeline(zip_it=True)).grid(row=4, column=1)
    tk.Button(root, text="Run All (Feed + ZIP + SFTP)", command=lambda: run_pipeline(zip_it=True, sftp_it=True)).grid(row=4, column=2, columnspan=2)

    # Log Output
    log_box = scrolledtext.ScrolledText(root, width=90, height=20)
    log_box.grid(row=5, column=0, columnspan=4, padx=10, pady=10)

    root.mainloop()

if __name__ == "__main__":
    launch_gui()
